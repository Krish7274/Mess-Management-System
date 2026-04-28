import os
import random
import string
import re
from datetime import datetime, timedelta

from werkzeug.utils import secure_filename
from flask import send_from_directory, Blueprint, request, jsonify
from flask_jwt_extended import (
    create_access_token, jwt_required, get_jwt, get_jwt_identity
)
from openpyxl import load_workbook

from models import (
    db,
    User,
    Bill,
    Payment,
    Notification,
    EmailOTP,
    Menu,
    WeeklyMenu,
    Attendance,
    Inventory,
    Complaint,
    HelpTicket,
    HelpMessage,
    MealPlan,
)
from utils.email_service import (
    send_bill_email,
    send_otp_email,
    send_user_welcome_email,
    send_meal_confirmation_email,
)

api = Blueprint("api", __name__)


@api.get("/uploads/<path:filename>")
def uploaded_file(filename):
    upload_dir = os.path.join(os.getcwd(), "uploads")
    return send_from_directory(upload_dir, filename)


def generate_temp_password(length=10):
    chars = string.ascii_letters + string.digits + "@#"
    return "".join(random.choice(chars) for _ in range(length))


def require_roles(*roles):
    def wrapper():
        claims = get_jwt()
        return claims.get("role") in roles
    return wrapper


def parse_flexible_date(date_str):
    date_str = (date_str or "").strip()
    if not date_str:
        raise ValueError("Empty date")

    for fmt in ("%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            pass

    raise ValueError(f"Invalid date format: {date_str}")


def normalize_to_iso_date(date_str):
    dt = parse_flexible_date(date_str)
    return dt.strftime("%Y-%m-%d")


def normalize_email(email):
    return str(email or "").strip().lower()


def normalize_text(value):
    return str(value or "").strip()


def is_valid_excel_file(filename):
    return bool(filename and filename.lower().endswith(".xlsx"))


def clean_contact(value):
    raw = str(value or "").strip()
    return re.sub(r"[^\d+]", "", raw)


def read_students_from_excel(file_storage):
    try:
        file_storage.stream.seek(0)
        workbook = load_workbook(file_storage.stream, data_only=True)
        sheet = workbook.active
    except Exception as e:
        print("EXCEL READ ERROR:", e)
        return [], "Unable to read Excel file. Please upload a valid .xlsx file."

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], "Excel file is empty"

    headers = [normalize_text(h).lower() for h in rows[0]]

    aliases = {
        "name": ["name", "student name", "full name"],
        "email": ["email", "mail", "mail-id", "mail id", "email id", "email-id"],
        "contact": ["contact", "mobile", "mob", "mobile number", "mob number", "phone", "phone number"],
        "room_no": ["room no", "room_no", "room", "room number"],
    }

    header_map = {
        "name": None,
        "email": None,
        "contact": None,
        "room_no": None,
    }

    for field, possible_names in aliases.items():
        for index, header in enumerate(headers):
            if header in possible_names:
                header_map[field] = index
                break

    if header_map["name"] is None or header_map["email"] is None:
        return [], "Excel must contain Name and Email columns"

    students = []

    for row_number, row in enumerate(rows[1:], start=2):
        if not row:
            continue

        def get_cell(col_index):
            if col_index is None:
                return ""
            if col_index >= len(row):
                return ""
            return row[col_index]

        name = normalize_text(get_cell(header_map["name"]))
        email = normalize_email(get_cell(header_map["email"]))
        contact = clean_contact(get_cell(header_map["contact"]))
        room_no = normalize_text(get_cell(header_map["room_no"]))

        if not name and not email and not contact and not room_no:
            continue

        students.append({
            "row_number": row_number,
            "name": name,
            "email": email,
            "contact": contact,
            "room_no": room_no,
        })

    return students, None


@api.post("/auth/send-otp")
def send_otp():
    try:
        data = request.get_json() or {}
        email = (data.get("email") or "").lower().strip()

        if not email:
            return jsonify({"error": "Email required"}), 400

        if User.query.filter_by(email=email).first():
            return jsonify({"error": "Email already exists"}), 409

        otp = str(random.randint(100000, 999999))
        expires = datetime.utcnow() + timedelta(minutes=10)

        EmailOTP.query.filter_by(email=email).delete()

        new_otp = EmailOTP(
            email=email,
            otp=otp,
            expires_at=expires
        )
        db.session.add(new_otp)
        db.session.commit()

        send_otp_email(email, otp)

        return jsonify({"message": "OTP sent successfully"}), 200

    except Exception as e:
        db.session.rollback()
        print("OTP ERROR:", e)
        return jsonify({"error": "Failed to send OTP"}), 500


@api.post("/auth/register")
def register():
    try:
        data = request.get_json() or {}

        name = (data.get("name") or "").strip()
        email = (data.get("email") or "").lower().strip()
        password = data.get("password")
        contact = data.get("contact")
        room_no = data.get("room_no")
        otp = (data.get("otp") or "").strip()

        if not name or not email or not password or not otp:
            return jsonify({"error": "All fields + OTP required"}), 400

        if User.query.filter_by(email=email).first():
            return jsonify({"error": "Email already exists"}), 409

        otp_row = EmailOTP.query.filter_by(email=email, otp=otp).first()

        if not otp_row:
            return jsonify({"error": "Invalid OTP"}), 400

        if datetime.utcnow() > otp_row.expires_at:
            return jsonify({"error": "OTP expired"}), 400

        user = User(
            name=name,
            email=email,
            contact=contact,
            room_no=room_no,
            role="User",
            must_change_password=False,
            password_changed_at=datetime.utcnow()
        )
        user.set_password(password)

        db.session.add(user)
        db.session.delete(otp_row)
        db.session.commit()

        return jsonify({"message": "Registered successfully"}), 201

    except Exception as e:
        db.session.rollback()
        print("REGISTER ERROR:", e)
        return jsonify({"error": "Registration failed"}), 500


@api.post("/auth/login")
def login():
    try:
        data = request.get_json() or {}
        email = (data.get("email") or "").lower().strip()
        password = data.get("password") or ""

        if not email or not password:
            return jsonify({"error": "email and password required"}), 400

        u = User.query.filter_by(email=email).first()
        if not u or not u.check_password(password):
            return jsonify({"error": "Invalid email or password"}), 401

        token = create_access_token(
            identity=str(u.id),
            additional_claims={
                "role": u.role,
                "name": u.name,
                "email": u.email,
                "must_change_password": u.must_change_password
            }
        )

        return jsonify({
            "access_token": token,
            "user": {
                "id": u.id,
                "name": u.name,
                "email": u.email,
                "role": u.role,
                "contact": u.contact,
                "room_no": u.room_no,
                "must_change_password": u.must_change_password
            }
        }), 200

    except Exception as e:
        print("LOGIN ERROR:", str(e))
        return jsonify({"error": "Login failed"}), 500


@api.post("/auth/change-password")
@jwt_required()
def change_password():
    try:
        uid = int(get_jwt_identity())
        user = User.query.get_or_404(uid)

        data = request.get_json() or {}
        current_password = (data.get("current_password") or "").strip()
        new_password = (data.get("new_password") or "").strip()
        confirm_password = (data.get("confirm_password") or "").strip()

        if not current_password or not new_password or not confirm_password:
            return jsonify({"error": "All password fields are required"}), 400

        if not user.check_password(current_password):
            return jsonify({"error": "Current password is incorrect"}), 400

        if len(new_password) < 6:
            return jsonify({"error": "New password must be at least 6 characters"}), 400

        if new_password != confirm_password:
            return jsonify({"error": "New password and confirm password do not match"}), 400

        if current_password == new_password:
            return jsonify({"error": "New password must be different from current password"}), 400

        user.set_password(new_password)
        user.must_change_password = False
        user.password_changed_at = datetime.utcnow()
        db.session.commit()

        new_token = create_access_token(
            identity=str(user.id),
            additional_claims={
                "role": user.role,
                "name": user.name,
                "email": user.email,
                "must_change_password": user.must_change_password
            }
        )

        return jsonify({
            "message": "Password changed successfully",
            "access_token": new_token,
            "user": {
                "id": user.id,
                "name": user.name,
                "email": user.email,
                "role": user.role,
                "contact": user.contact,
                "room_no": user.room_no,
                "must_change_password": user.must_change_password
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        print("CHANGE PASSWORD ERROR:", e)
        return jsonify({"error": "Failed to change password"}), 500


@api.get("/users/me")
@jwt_required()
def me():
    uid = int(get_jwt_identity())
    u = User.query.get_or_404(uid)

    return jsonify({
        "id": u.id,
        "name": u.name,
        "email": u.email,
        "role": u.role,
        "contact": u.contact,
        "room_no": u.room_no,
        "must_change_password": u.must_change_password
    })


@api.put("/users/me")
@jwt_required()
def update_me():
    try:
        uid = int(get_jwt_identity())
        user = User.query.get_or_404(uid)

        data = request.get_json() or {}

        name = (data.get("name") or "").strip()
        contact = (data.get("contact") or "").strip()
        room_no = (data.get("room_no") or "").strip()

        if not name:
            return jsonify({"error": "Name is required"}), 400

        user.name = name
        user.contact = contact
        user.room_no = room_no

        db.session.commit()

        return jsonify({
            "message": "Profile updated successfully",
            "user": {
                "id": user.id,
                "name": user.name,
                "email": user.email,
                "role": user.role,
                "contact": user.contact,
                "room_no": user.room_no,
                "must_change_password": user.must_change_password
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        print("UPDATE PROFILE ERROR:", e)
        return jsonify({"error": "Failed to update profile"}), 500


@api.post("/users")
@jwt_required()
def add_user():
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        data = request.get_json() or {}

        name = (data.get("name") or "").strip()
        email = (data.get("email") or "").lower().strip()
        contact = (data.get("contact") or "").strip()
        room_no = (data.get("room_no") or "").strip()

        if not name:
            return jsonify({"error": "Name is required"}), 400
        if not email:
            return jsonify({"error": "Email is required"}), 400

        creator_role = get_jwt().get("role")
        new_role = (data.get("role") or "User").strip()

        if creator_role == "Staff":
            new_role = "User"

        if creator_role == "Admin" and new_role not in ["Admin", "Staff", "User"]:
            return jsonify({"error": "Role must be Admin, Staff, or User"}), 400

        existing = User.query.filter_by(email=email).first()
        if existing:
            return jsonify({"error": "Email already exists"}), 400

        temp_password = generate_temp_password()

        new_user = User(
            name=name,
            email=email,
            role=new_role,
            contact=contact,
            room_no=room_no,
            must_change_password=True
        )
        new_user.set_password(temp_password)

        db.session.add(new_user)
        db.session.commit()

        email_status = "sent"
        email_error = None
        try:
            send_user_welcome_email(new_user.email, new_user.name, temp_password)
        except Exception as mail_error:
            email_status = "failed"
            email_error = str(mail_error)
            print("WELCOME MAIL ERROR:", mail_error)

        return jsonify({
            "message": "User added successfully",
            "email_status": email_status,
            "email_error": email_error,
            "temp_password": temp_password,
            "user": {
                "id": new_user.id,
                "name": new_user.name,
                "email": new_user.email,
                "role": new_user.role,
                "contact": new_user.contact,
                "room_no": new_user.room_no,
                "must_change_password": new_user.must_change_password
            }
        }), 201

    except Exception as e:
        db.session.rollback()
        print("ADD USER ERROR:", e)
        return jsonify({"error": "Failed to add user"}), 500


@api.post("/users/import")
@jwt_required()
def import_users_from_excel():
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        if "file" not in request.files:
            return jsonify({"error": "Excel file is required"}), 400

        file = request.files["file"]

        if not file or not file.filename:
            return jsonify({"error": "Please choose an Excel file"}), 400

        if not is_valid_excel_file(file.filename):
            return jsonify({"error": "Only .xlsx Excel files are allowed"}), 400

        students, read_error = read_students_from_excel(file)
        if read_error:
            return jsonify({"error": read_error}), 400

        if not students:
            return jsonify({"error": "No valid student rows found in Excel"}), 400

        added_users = []
        skipped_users = []
        failed_users = []

        for student in students:
            row_number = student["row_number"]
            name = student["name"]
            email = student["email"]
            contact = student["contact"]
            room_no = student["room_no"]

            try:
                if not name:
                    skipped_users.append({
                        "row_number": row_number,
                        "email": email or "-",
                        "reason": "Name is required"
                    })
                    continue

                if not email:
                    skipped_users.append({
                        "row_number": row_number,
                        "email": "-",
                        "reason": "Email is required"
                    })
                    continue

                existing = User.query.filter_by(email=email).first()
                if existing:
                    skipped_users.append({
                        "row_number": row_number,
                        "email": email,
                        "reason": "Email already exists"
                    })
                    continue

                temp_password = generate_temp_password()

                new_user = User(
                    name=name,
                    email=email,
                    role="User",
                    contact=contact,
                    room_no=room_no,
                    must_change_password=True
                )
                new_user.set_password(temp_password)

                db.session.add(new_user)
                db.session.commit()

                email_status = "sent"
                email_error = None
                try:
                    send_user_welcome_email(new_user.email, new_user.name, temp_password)
                except Exception as mail_error:
                    email_status = "failed"
                    email_error = str(mail_error)
                    print(f"IMPORT MAIL ERROR for {email}:", mail_error)

                added_users.append({
                    "id": new_user.id,
                    "row_number": row_number,
                    "name": new_user.name,
                    "email": new_user.email,
                    "role": new_user.role,
                    "contact": new_user.contact,
                    "room_no": new_user.room_no,
                    "temp_password": temp_password,
                    "email_status": email_status,
                    "email_error": email_error
                })

            except Exception as row_error:
                db.session.rollback()
                print(f"IMPORT ROW ERROR row {row_number}:", row_error)
                failed_users.append({
                    "row_number": row_number,
                    "email": email or "-",
                    "reason": str(row_error)
                })

        sent_count = len([u for u in added_users if u["email_status"] == "sent"])
        failed_mail_count = len([u for u in added_users if u["email_status"] == "failed"])

        message = (
            f"Excel import completed. Added {len(added_users)} users, "
            f"skipped {len(skipped_users)}, failed rows {len(failed_users)}, "
            f"email sent {sent_count}, email failed {failed_mail_count}."
        )

        return jsonify({
            "message": message,
            "summary": {
                "total_rows": len(students),
                "added_count": len(added_users),
                "skipped_count": len(skipped_users),
                "failed_count": len(failed_users),
                "email_sent_count": sent_count,
                "email_failed_count": failed_mail_count
            },
            "added_users": added_users,
            "skipped_users": skipped_users,
            "failed_users": failed_users
        }), 200

    except Exception as e:
        db.session.rollback()
        print("IMPORT USERS ERROR:", e)
        return jsonify({"error": f"Failed to import users from Excel: {str(e)}"}), 500


@api.get("/users")
@jwt_required()
def list_users():
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    users = User.query.order_by(User.id.desc()).all()

    return jsonify([
        {
            "id": u.id,
            "name": u.name,
            "email": u.email,
            "role": u.role,
            "contact": u.contact,
            "room_no": u.room_no,
            "must_change_password": u.must_change_password
        }
        for u in users
    ]), 200


@api.put("/users/<int:user_id>/role")
@jwt_required()
def update_user_role(user_id):
    if not require_roles("Admin")():
        return jsonify({"error": "Only admin can change roles"}), 403

    try:
        current_user = User.query.get_or_404(int(get_jwt_identity()))
        user = User.query.get_or_404(user_id)

        data = request.get_json() or {}
        new_role = (data.get("role") or "").strip()

        if new_role not in ["Admin", "User", "Staff"]:
            return jsonify({"error": "role must be Admin, User, or Staff"}), 400

        if current_user.id == user.id and new_role != "Admin":
            return jsonify({"error": "Admin cannot remove own admin role"}), 400

        user.role = new_role
        db.session.commit()

        return jsonify({
            "message": f"{user.name} role updated to {new_role}",
            "user": {
                "id": user.id,
                "name": user.name,
                "email": user.email,
                "role": user.role,
                "contact": user.contact,
                "room_no": user.room_no,
                "must_change_password": user.must_change_password
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        print("UPDATE ROLE ERROR:", e)
        return jsonify({"error": "Failed to update role"}), 500


@api.delete("/users/<int:user_id>")
@jwt_required()
def delete_user(user_id):
    if not require_roles("Admin")():
        return jsonify({"error": "Only admin can delete users"}), 403

    try:
        current_user = User.query.get_or_404(int(get_jwt_identity()))
        user = User.query.get_or_404(user_id)

        if current_user.id == user.id:
            return jsonify({"error": "Admin cannot delete own account"}), 400

        if user.role == "Admin":
            return jsonify({"error": "Admin account cannot be deleted"}), 400

        user_bills = Bill.query.filter_by(user_id=user.id).all()
        bill_ids = [b.id for b in user_bills]

        if bill_ids:
            Payment.query.filter(Payment.bill_id.in_(bill_ids)).delete(synchronize_session=False)

        Bill.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        Attendance.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        MealPlan.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        Complaint.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        Notification.query.filter_by(user_id=user.id).delete(synchronize_session=False)
        EmailOTP.query.filter_by(email=user.email).delete(synchronize_session=False)

        deleted_name = user.name
        deleted_role = user.role

        db.session.delete(user)
        db.session.commit()

        return jsonify({
            "message": f"{deleted_role} '{deleted_name}' deleted successfully"
        }), 200

    except Exception as e:
        db.session.rollback()
        print("DELETE USER ERROR:", e)
        return jsonify({"error": "Failed to delete user"}), 500


@api.get("/menu")
@jwt_required()
def get_menu():
    items = Menu.query.order_by(Menu.id.desc()).all()
    return jsonify([
        {
            "id": m.id,
            "date": m.date,
            "meal_type": m.meal_type,
            "items": m.items
        }
        for m in items
    ])


@api.post("/menu")
@jwt_required()
def add_menu():
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        data = request.get_json() or {}

        date = normalize_to_iso_date(data.get("date"))
        meal_type = (data.get("meal_type") or "").strip()
        items = (data.get("items") or "").strip()

        if not date or not meal_type or not items:
            return jsonify({"error": "date, meal_type and items are required"}), 400

        if meal_type not in ["Breakfast", "Lunch", "Dinner"]:
            return jsonify({"error": "meal_type must be Breakfast/Lunch/Dinner"}), 400

        existing = Menu.query.filter_by(date=date, meal_type=meal_type).first()

        if existing:
            existing.items = items
            db.session.commit()
            return jsonify({"message": "Menu updated successfully"}), 200

        menu = Menu(date=date, meal_type=meal_type, items=items)
        db.session.add(menu)
        db.session.commit()

        return jsonify({"message": "Menu added successfully"}), 201

    except Exception as e:
        db.session.rollback()
        print("MENU ERROR:", e)
        return jsonify({"error": "Failed to save menu"}), 500


@api.delete("/menu/<int:menu_id>")
@jwt_required()
def delete_menu(menu_id):
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        menu = Menu.query.get_or_404(menu_id)
        db.session.delete(menu)
        db.session.commit()
        return jsonify({"message": "Menu deleted successfully"}), 200
    except Exception as e:
        db.session.rollback()
        print("DELETE MENU ERROR:", e)
        return jsonify({"error": "Failed to delete menu"}), 500


def default_weekly_structure():
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    meals = ["Breakfast", "Lunch", "Dinner"]

    data = {}
    for day in days:
        data[day] = {}
        for meal in meals:
            data[day][meal] = {
                "items": "",
                "price": 0
            }
    return data


def normalize_weekly_items(raw):
    base = default_weekly_structure()
    raw = raw or {}

    for day, meals in base.items():
        incoming_day = raw.get(day, {})
        for meal in meals:
            incoming_meal = incoming_day.get(meal, {})
            items = (incoming_meal.get("items") or "").strip()

            try:
                price = float(incoming_meal.get("price", 0) or 0)
            except Exception:
                price = 0

            if price < 0:
                price = 0

            base[day][meal] = {
                "items": items,
                "price": round(price, 2)
            }

    return base


def serialize_weekly_menu(menu):
    return {
        "id": menu.id,
        "week_start": menu.week_start,
        "weekly_items": menu.weekly_items,
        "created_at": menu.created_at.isoformat() if menu.created_at else None,
        "updated_at": menu.updated_at.isoformat() if menu.updated_at else None
    }


def get_week_start_from_date(date_str):
    dt = parse_flexible_date(date_str)
    day = dt.weekday()
    monday = dt - timedelta(days=day)
    return monday.strftime("%Y-%m-%d")


def get_day_name_from_date(date_str):
    dt = parse_flexible_date(date_str)
    return dt.strftime("%A")


def get_today_menu_from_weekly_menu(date_str):
    week_start = get_week_start_from_date(date_str)
    day_name = get_day_name_from_date(date_str)

    weekly_menu = WeeklyMenu.query.filter_by(week_start=week_start).first()
    if not weekly_menu:
        return None, f"No weekly menu found for week starting {week_start}"

    weekly_items = weekly_menu.weekly_items or {}
    day_block = weekly_items.get(day_name, {}) or {}

    return {
        "week_start": week_start,
        "date": normalize_to_iso_date(date_str),
        "day": day_name,
        "breakfast": day_block.get("Breakfast", {"items": "", "price": 0}),
        "lunch": day_block.get("Lunch", {"items": "", "price": 0}),
        "dinner": day_block.get("Dinner", {"items": "", "price": 0}),
    }, None


def get_meal_price_from_weekly_menu(date_str, meal_type):
    week_start = get_week_start_from_date(date_str)
    day_name = get_day_name_from_date(date_str)

    weekly_menu = WeeklyMenu.query.filter_by(week_start=week_start).first()
    if not weekly_menu:
        return None, f"No weekly menu found for week starting {week_start}"

    weekly_items = weekly_menu.weekly_items or {}
    day_block = weekly_items.get(day_name, {})
    meal_block = day_block.get(meal_type, {})

    try:
        price = float(meal_block.get("price", 0) or 0)
    except Exception:
        price = 0

    if price <= 0:
        return None, f"No valid price found for {day_name} {meal_type}"

    return round(price, 2), None


def bill_belongs_to_month(bill_row, month_value):
    return str(bill_row.month or "").startswith(month_value)


def recalculate_parent_monthly_bill(parent_bill_id):
    parent_bill = Bill.query.get(parent_bill_id)
    if not parent_bill:
        return None

    child_bills = Bill.query.filter_by(parent_bill_id=parent_bill.id).all()
    active_children = [b for b in child_bills if b.status in ["Merged", "Paid"]]

    if not active_children and parent_bill.status != "Paid":
        db.session.delete(parent_bill)
        return None

    parent_bill.amount = round(sum(float(b.amount or 0) for b in active_children), 2)
    return parent_bill


def get_unmerged_daily_bills_for_month(user_id, month_value):
    daily_bills = Bill.query.filter_by(user_id=user_id, bill_type="daily").all()
    return [
        b for b in daily_bills
        if bill_belongs_to_month(b, month_value)
        and b.parent_bill_id is None
        and b.status == "Unpaid"
    ]


def serialize_bill_row(bill_row, include_user=False):
    payment = Payment.query.filter_by(bill_id=bill_row.id).order_by(Payment.id.desc()).first()
    child_count = Bill.query.filter_by(parent_bill_id=bill_row.id).count()
    user = User.query.get(bill_row.user_id) if include_user else None

    data = {
        "id": bill_row.id,
        "user_id": bill_row.user_id,
        "bill_type": bill_row.bill_type,
        "meal_type": bill_row.meal_type,
        "attendance_id": bill_row.attendance_id,
        "parent_bill_id": bill_row.parent_bill_id,
        "period": bill_row.month,
        "amount": round(float(bill_row.amount or 0), 2),
        "status": bill_row.status,
        "included_meals_count": child_count,
        "can_pay": bill_row.status == "Unpaid" and bill_row.parent_bill_id is None,
        "payment": {
            "mode": payment.mode,
            "receipt_no": payment.receipt_no,
            "proof_url": f"http://127.0.0.1:5000/api/uploads/{payment.proof_filename}" if payment and payment.proof_filename else None,
            "paid_at": payment.paid_at.isoformat() if payment and payment.paid_at else None
        } if payment else None
    }

    if include_user:
        data["user_name"] = user.name if user else "Unknown"
        data["user_email"] = user.email if user else "Unknown"

    return data


def sync_attendance_bill(attendance_row):
    existing_bill = Bill.query.filter_by(attendance_id=attendance_row.id).first()

    if attendance_row.status == "Taken":
        meal_price, error = get_meal_price_from_weekly_menu(
            attendance_row.date,
            attendance_row.meal_type
        )
        if error:
            return False, error, None

        if existing_bill:
            if existing_bill.parent_bill_id:
                parent_bill = Bill.query.get(existing_bill.parent_bill_id)
                if parent_bill and parent_bill.status == "Paid":
                    return False, "This attendance is already settled inside a paid monthly bill", None

                existing_bill.user_id = attendance_row.user_id
                existing_bill.month = attendance_row.date
                existing_bill.bill_type = "daily"
                existing_bill.meal_type = attendance_row.meal_type
                existing_bill.attendance_id = attendance_row.id
                existing_bill.amount = meal_price
                existing_bill.status = "Merged"

                if parent_bill:
                    recalculate_parent_monthly_bill(parent_bill.id)

                return True, "Attendance bill updated and monthly total refreshed", existing_bill

            if existing_bill.status == "Paid":
                return False, "This attendance bill is already paid and cannot be changed", None

            existing_bill.user_id = attendance_row.user_id
            existing_bill.month = attendance_row.date
            existing_bill.bill_type = "daily"
            existing_bill.meal_type = attendance_row.meal_type
            existing_bill.attendance_id = attendance_row.id
            existing_bill.amount = meal_price
            existing_bill.status = "Unpaid"
            return True, "Attendance bill updated", existing_bill

        new_bill = Bill(
            user_id=attendance_row.user_id,
            month=attendance_row.date,
            bill_type="daily",
            meal_type=attendance_row.meal_type,
            attendance_id=attendance_row.id,
            parent_bill_id=None,
            amount=meal_price,
            status="Unpaid"
        )
        db.session.add(new_bill)
        db.session.flush()

        user = User.query.get(attendance_row.user_id)
        if user:
            notification = Notification(
                user_id=user.id,
                title="Meal Bill Generated",
                message=f"{attendance_row.meal_type} bill for {attendance_row.date} generated: ₹{meal_price}"
            )
            db.session.add(notification)

            if user.email:
                try:
                    send_bill_email(
                        user.email,
                        user.name,
                        f"{attendance_row.meal_type} Meal",
                        attendance_row.date,
                        meal_price
                    )
                except Exception as mail_error:
                    print("AUTO ATTENDANCE BILL EMAIL ERROR:", mail_error)

        return True, "Attendance bill created", new_bill

    if existing_bill:
        if existing_bill.parent_bill_id:
            parent_bill = Bill.query.get(existing_bill.parent_bill_id)
            if parent_bill and parent_bill.status == "Paid":
                return False, "This attendance is already settled inside a paid monthly bill", None
            db.session.delete(existing_bill)
            if parent_bill:
                recalculate_parent_monthly_bill(parent_bill.id)
            return True, "Attendance removed and monthly total refreshed", None

        if existing_bill.status == "Paid":
            return False, "Bill already paid, cannot auto-delete it after skipping attendance", None

        Payment.query.filter_by(bill_id=existing_bill.id).delete()
        db.session.delete(existing_bill)

    return True, "Attendance bill removed", None


@api.get("/menu/weekly")
@jwt_required()
def get_weekly_menus():
    menus = WeeklyMenu.query.order_by(WeeklyMenu.week_start.desc()).all()
    return jsonify([serialize_weekly_menu(m) for m in menus])


@api.get("/menu/weekly/today")
@jwt_required()
def get_today_weekly_menu():
    try:
        requested_date = request.args.get("date")

        if requested_date:
            date_str = normalize_to_iso_date(requested_date)
        else:
            date_str = datetime.utcnow().strftime("%Y-%m-%d")

        data, error = get_today_menu_from_weekly_menu(date_str)

        if error:
            return jsonify({"error": error}), 404

        return jsonify(data), 200

    except Exception as e:
        print("GET TODAY WEEKLY MENU ERROR:", e)
        return jsonify({"error": "Failed to load today's weekly menu"}), 500


@api.post("/menu/weekly")
@jwt_required()
def create_weekly_menu():
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        data = request.get_json() or {}
        week_start = normalize_to_iso_date(data.get("week_start"))
        weekly_items = data.get("weekly_items") or {}

        if not week_start:
            return jsonify({"error": "week_start is required"}), 400

        normalized = normalize_weekly_items(weekly_items)

        existing = WeeklyMenu.query.filter_by(week_start=week_start).first()
        if existing:
            existing.weekly_items = normalized
            db.session.commit()
            return jsonify({
                "message": "Weekly menu updated successfully",
                "menu": serialize_weekly_menu(existing)
            }), 200

        menu = WeeklyMenu(
            week_start=week_start,
            weekly_items=normalized
        )
        db.session.add(menu)
        db.session.commit()

        return jsonify({
            "message": "Weekly menu created successfully",
            "menu": serialize_weekly_menu(menu)
        }), 201

    except Exception as e:
        db.session.rollback()
        print("CREATE WEEKLY MENU ERROR:", e)
        return jsonify({"error": "Failed to create weekly menu"}), 500


@api.put("/menu/weekly/<int:menu_id>")
@jwt_required()
def update_weekly_menu(menu_id):
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        menu = WeeklyMenu.query.get_or_404(menu_id)
        data = request.get_json() or {}

        week_start = normalize_to_iso_date(data.get("week_start"))
        weekly_items = data.get("weekly_items") or {}

        if not week_start:
            return jsonify({"error": "week_start is required"}), 400

        duplicate = WeeklyMenu.query.filter(
            WeeklyMenu.week_start == week_start,
            WeeklyMenu.id != menu_id
        ).first()
        if duplicate:
            return jsonify({"error": "Another weekly menu already exists for this week start"}), 400

        menu.week_start = week_start
        menu.weekly_items = normalize_weekly_items(weekly_items)
        db.session.commit()

        return jsonify({
            "message": "Weekly menu updated successfully",
            "menu": serialize_weekly_menu(menu)
        }), 200

    except Exception as e:
        db.session.rollback()
        print("UPDATE WEEKLY MENU ERROR:", e)
        return jsonify({"error": "Failed to update weekly menu"}), 500


@api.delete("/menu/weekly/<int:menu_id>")
@jwt_required()
def delete_weekly_menu(menu_id):
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        menu = WeeklyMenu.query.get_or_404(menu_id)
        db.session.delete(menu)
        db.session.commit()
        return jsonify({"message": "Weekly menu deleted successfully"}), 200
    except Exception as e:
        db.session.rollback()
        print("DELETE WEEKLY MENU ERROR:", e)
        return jsonify({"error": "Failed to delete weekly menu"}), 500


@api.get("/attendance")
@jwt_required()
def attendance_list():
    role = get_jwt().get("role")
    uid = int(get_jwt_identity())

    date = request.args.get("date")
    meal_type = request.args.get("meal_type")

    q = Attendance.query

    if role not in ["Admin", "Staff"]:
        q = q.filter_by(user_id=uid)

    if date:
        try:
            date = normalize_to_iso_date(date)
        except Exception:
            return jsonify({"error": "Invalid date format"}), 400
        q = q.filter_by(date=date)

    if meal_type:
        q = q.filter_by(meal_type=meal_type)

    items = q.order_by(Attendance.id.desc()).all()

    return jsonify([
        {
            "id": a.id,
            "user_id": a.user_id,
            "date": a.date,
            "meal_type": a.meal_type,
            "status": a.status
        }
        for a in items
    ])


@api.post("/attendance")
@jwt_required()
def mark_attendance():
    try:
        data = request.get_json() or {}

        role = get_jwt().get("role")
        logged_in_uid = int(get_jwt_identity())

        date = normalize_to_iso_date(data.get("date"))
        meal_type = (data.get("meal_type") or "").strip()
        status = (data.get("status") or "").strip()

        if not date or not meal_type or not status:
            return jsonify({"error": "date, meal_type, status required"}), 400

        if meal_type not in ["Breakfast", "Lunch", "Dinner"]:
            return jsonify({"error": "meal_type must be Breakfast/Lunch/Dinner"}), 400

        if status not in ["Taken", "Skipped"]:
            return jsonify({"error": "status must be Taken/Skipped"}), 400

        target_user_id = logged_in_uid

        if role in ["Admin", "Staff"] and data.get("user_id") is not None:
            try:
                target_user_id = int(data.get("user_id"))
            except Exception:
                return jsonify({"error": "user_id must be a number"}), 400

        existing = Attendance.query.filter_by(
            user_id=target_user_id,
            date=date,
            meal_type=meal_type
        ).first()

        if existing:
            existing.status = status
            ok, billing_message, bill_row = sync_attendance_bill(existing)
            if not ok:
                db.session.rollback()
                return jsonify({"error": billing_message}), 400

            db.session.commit()
            return jsonify({
                "message": "Attendance updated successfully",
                "billing_message": billing_message,
                "bill": serialize_bill_row(bill_row) if bill_row else None
            }), 200

        new_attendance = Attendance(
            user_id=target_user_id,
            date=date,
            meal_type=meal_type,
            status=status
        )
        db.session.add(new_attendance)
        db.session.flush()

        ok, billing_message, bill_row = sync_attendance_bill(new_attendance)
        if not ok:
            db.session.rollback()
            return jsonify({"error": billing_message}), 400

        db.session.commit()

        return jsonify({
            "message": "Attendance saved successfully",
            "billing_message": billing_message,
            "bill": serialize_bill_row(bill_row) if bill_row else None
        }), 201

    except Exception as e:
        db.session.rollback()
        print("ATTENDANCE ERROR:", e)
        return jsonify({"error": "Failed to save attendance"}), 500

def formatDateForMail(date_str):
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    except Exception:
        return date_str


def get_meal_plan_price_summary(plan_date, breakfast, lunch, dinner):
    selected_meals = []
    total_amount = 0

    requested_meals = []

    if breakfast:
        requested_meals.append("Breakfast")
    if lunch:
        requested_meals.append("Lunch")
    if dinner:
        requested_meals.append("Dinner")

    for meal in requested_meals:
        price, error = get_meal_price_from_weekly_menu(plan_date, meal)

        if error:
            price = 0

        selected_meals.append({
            "name": meal,
            "price": round(float(price or 0), 2)
        })

        total_amount += float(price or 0)

    return selected_meals, round(total_amount, 2)

def serialize_meal_plan(plan, include_user=False):
    user = User.query.get(plan.user_id) if include_user else None
    return {
        "id": plan.id,
        "user_id": plan.user_id,
        "user_name": user.name if user else None,
        "email": user.email if user else None,
        "date": plan.date,
        "breakfast": bool(plan.breakfast),
        "lunch": bool(plan.lunch),
        "dinner": bool(plan.dinner),
        "created_at": plan.created_at.isoformat() if plan.created_at else None,
        "updated_at": plan.updated_at.isoformat() if plan.updated_at else None,
    }


@api.get("/meal-plans/my")
@jwt_required()
def get_my_meal_plan():
    try:
        uid = int(get_jwt_identity())
        date = normalize_to_iso_date(
            request.args.get("date") or datetime.utcnow().strftime("%Y-%m-%d")
        )

        plan = MealPlan.query.filter_by(user_id=uid, date=date).first()
        if not plan:
            return jsonify({
                "date": date,
                "breakfast": False,
                "lunch": False,
                "dinner": False,
            }), 200

        return jsonify(serialize_meal_plan(plan)), 200

    except Exception as e:
        print("GET MY MEAL PLAN ERROR:", e)
        return jsonify({"error": "Failed to load meal plan"}), 500


@api.get("/meal-plans")
@jwt_required()
def get_all_meal_plans():
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        date = normalize_to_iso_date(
            request.args.get("date") or datetime.utcnow().strftime("%Y-%m-%d")
        )
        plans = MealPlan.query.filter_by(date=date).order_by(MealPlan.id.desc()).all()
        return jsonify([serialize_meal_plan(p, include_user=True) for p in plans]), 200

    except Exception as e:
        print("GET ALL MEAL PLANS ERROR:", e)
        return jsonify({"error": "Failed to load meal confirmations"}), 500


@api.post("/meal-plans")
@jwt_required()
def save_meal_plan():
    try:
        uid = int(get_jwt_identity())
        user = User.query.get_or_404(uid)

        data = request.get_json() or {}

        plan_date = normalize_to_iso_date(data.get("date"))
        breakfast = bool(data.get("breakfast"))
        lunch = bool(data.get("lunch"))
        dinner = bool(data.get("dinner"))

        existing = MealPlan.query.filter_by(
            user_id=user.id,
            date=plan_date
        ).first()

        if existing:
            existing.breakfast = breakfast
            existing.lunch = lunch
            existing.dinner = dinner
            existing.updated_at = datetime.utcnow()
            meal_plan = existing
        else:
            meal_plan = MealPlan(
                user_id=user.id,
                date=plan_date,
                breakfast=breakfast,
                lunch=lunch,
                dinner=dinner
            )
            db.session.add(meal_plan)
            db.session.flush()

        selected_meals, total_amount = get_meal_plan_price_summary(
            plan_date,
            breakfast,
            lunch,
            dinner
        )

        db.session.add(Notification(
            user_id=user.id,
            title="Meal Confirmation Saved",
            message=f"Your meal confirmation for {formatDateForMail(plan_date)} has been saved. Estimated amount: ₹{total_amount}"
        ))

        email_sent = False
        email_error = None

        if user.email:
            try:
                send_meal_confirmation_email(
                    user.email,
                    user.name,
                    formatDateForMail(plan_date),
                    selected_meals,
                    total_amount
                )
                email_sent = True
            except Exception as mail_error:
                email_error = str(mail_error)
                print("MEAL CONFIRMATION EMAIL ERROR:", mail_error)

        db.session.commit()

        return jsonify({
            "message": f"Meal plan for {formatDateForMail(plan_date)} saved successfully. Estimated amount: ₹{total_amount}",
            "email_sent": email_sent,
            "email_error": email_error,
            "meal_plan": {
                "id": meal_plan.id,
                "user_id": meal_plan.user_id,
                "date": meal_plan.date,
                "breakfast": meal_plan.breakfast,
                "lunch": meal_plan.lunch,
                "dinner": meal_plan.dinner,
                "selected_meals": selected_meals,
                "total_amount": total_amount
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        print("SAVE MEAL PLAN ERROR:", e)
        return jsonify({"error": "Failed to save meal confirmation"}), 500

# =========================
# INVENTORY HELPER FUNCTIONS
# =========================

def inventory_status(item):
    qty = float(item.qty or 0)
    low_limit = float(item.low_limit or 0)

    if qty <= 0:
        return "Out of Stock"

    if qty <= low_limit:
        return "Low Stock"

    return "In Stock"


def serialize_inventory(item):
    qty = float(item.qty or 0)
    price_per_unit = float(item.price_per_unit or 0)

    return {
        "id": item.id,
        "category": item.category,
        "name": item.name,
        "unit": item.unit,
        "qty": qty,
        "low_limit": float(item.low_limit or 0),
        "price_per_unit": price_per_unit,
        "total_value": round(qty * price_per_unit, 2),
        "status": inventory_status(item),
        "low": qty > 0 and qty <= float(item.low_limit or 0),
        "updated_at": item.updated_at.isoformat() if item.updated_at else None
    }

@api.get("/inventory")
@jwt_required()
def get_inventory():
    items = Inventory.query.order_by(Inventory.id.desc()).all()
    return jsonify([serialize_inventory(i) for i in items])


@api.get("/inventory/summary")
@jwt_required()
def inventory_summary():
    items = Inventory.query.all()

    total_items = len(items)
    total_quantity = round(sum(float(i.qty or 0) for i in items), 2)
    low_stock = sum(1 for i in items if inventory_status(i) == "Low Stock")
    out_of_stock = sum(1 for i in items if inventory_status(i) == "Out of Stock")
    total_value = round(
        sum(float(i.qty or 0) * float(i.price_per_unit or 0) for i in items),
        2
    )

    return jsonify({
        "total_items": total_items,
        "total_quantity": total_quantity,
        "low_stock": low_stock,
        "out_of_stock": out_of_stock,
        "total_value": total_value
    })


@api.post("/inventory")
@jwt_required()
def add_inventory():
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        data = request.get_json() or {}

        category = (data.get("category") or "").strip()
        name = (data.get("name") or "").strip()
        unit = (data.get("unit") or "kg").strip()

        if not category or not name:
            return jsonify({"error": "category and name required"}), 400

        try:
            qty = float(data.get("qty", 0) or 0)
            low_limit = float(data.get("low_limit", 5) or 5)
            price_per_unit = float(data.get("price_per_unit", 0) or 0)
        except Exception:
            return jsonify({"error": "qty, low_limit and price_per_unit must be numbers"}), 400

        if qty < 0 or low_limit < 0 or price_per_unit < 0:
            return jsonify({"error": "qty, low_limit and price_per_unit cannot be negative"}), 400

        existing = Inventory.query.filter(
            db.func.lower(Inventory.category) == category.lower(),
            db.func.lower(Inventory.name) == name.lower()
        ).first()

        if existing:
            existing.unit = unit
            existing.qty = qty
            existing.low_limit = low_limit
            existing.price_per_unit = price_per_unit
            db.session.commit()
            return jsonify({
                "message": "Inventory item updated successfully",
                "item": serialize_inventory(existing)
            }), 200

        item = Inventory(
            category=category,
            name=name,
            unit=unit,
            qty=qty,
            low_limit=low_limit,
            price_per_unit=price_per_unit
        )
        db.session.add(item)
        db.session.commit()

        return jsonify({
            "message": "Inventory added successfully",
            "item": serialize_inventory(item)
        }), 201

    except Exception as e:
        db.session.rollback()
        print("INVENTORY ADD ERROR:", e)
        return jsonify({"error": "Failed to add inventory"}), 500


@api.put("/inventory/<int:item_id>")
@jwt_required()
def update_inventory(item_id):
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        item = Inventory.query.get_or_404(item_id)
        data = request.get_json() or {}

        category = (data.get("category") or "").strip()
        name = (data.get("name") or "").strip()
        unit = (data.get("unit") or "kg").strip()

        if not category or not name:
            return jsonify({"error": "category and name required"}), 400

        try:
            qty = float(data.get("qty", 0) or 0)
            low_limit = float(data.get("low_limit", 5) or 5)
            price_per_unit = float(data.get("price_per_unit", 0) or 0)
        except Exception:
            return jsonify({"error": "qty, low_limit and price_per_unit must be numbers"}), 400

        if qty < 0 or low_limit < 0 or price_per_unit < 0:
            return jsonify({"error": "qty, low_limit and price_per_unit cannot be negative"}), 400

        item.category = category
        item.name = name
        item.unit = unit
        item.qty = qty
        item.low_limit = low_limit
        item.price_per_unit = price_per_unit

        db.session.commit()

        return jsonify({
            "message": "Inventory updated successfully",
            "item": serialize_inventory(item)
        }), 200

    except Exception as e:
        db.session.rollback()
        print("INVENTORY UPDATE ERROR:", e)
        return jsonify({"error": "Failed to update inventory"}), 500


@api.patch("/inventory/<int:item_id>/adjust")
@jwt_required()
def adjust_inventory(item_id):
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        item = Inventory.query.get_or_404(item_id)
        data = request.get_json() or {}

        try:
            qty_delta = float(data.get("qty_delta", 0) or 0)
        except Exception:
            return jsonify({"error": "qty_delta must be a number"}), 400

        new_qty = float(item.qty or 0) + qty_delta
        if new_qty < 0:
            new_qty = 0

        item.qty = new_qty
        db.session.commit()

        return jsonify({
            "message": "Inventory quantity updated successfully",
            "item": serialize_inventory(item)
        }), 200

    except Exception as e:
        db.session.rollback()
        print("INVENTORY ADJUST ERROR:", e)
        return jsonify({"error": "Failed to update inventory quantity"}), 500


@api.delete("/inventory/<int:item_id>")
@jwt_required()
def delete_inventory(item_id):
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        item = Inventory.query.get_or_404(item_id)
        db.session.delete(item)
        db.session.commit()

        return jsonify({"message": "Inventory deleted successfully"}), 200

    except Exception as e:
        db.session.rollback()
        print("INVENTORY DELETE ERROR:", e)
        return jsonify({"error": "Failed to delete inventory"}), 500




def serialize_help_message(msg):
    user = User.query.get(msg.sender_id)
    return {
        "id": msg.id,
        "ticket_id": msg.ticket_id,
        "sender_id": msg.sender_id,
        "sender_name": user.name if user else "Unknown",
        "sender_role": msg.sender_role,
        "message": msg.message,
        "created_at": msg.created_at.isoformat() if msg.created_at else None,
    }


def serialize_help_ticket(ticket, include_messages=False):
    user = User.query.get(ticket.user_id)

    data = {
        "id": ticket.id,
        "user_id": ticket.user_id,
        "user_name": user.name if user else "Unknown",
        "user_email": user.email if user else "Unknown",
        "subject": ticket.subject,
        "category": ticket.category,
        "status": ticket.status,
        "created_at": ticket.created_at.isoformat() if ticket.created_at else None,
        "updated_at": ticket.updated_at.isoformat() if ticket.updated_at else None,
    }

    if include_messages:
        messages = HelpMessage.query.filter_by(ticket_id=ticket.id).order_by(HelpMessage.id.asc()).all()
        data["messages"] = [serialize_help_message(m) for m in messages]

    return data


@api.post("/help-centre/tickets")
@jwt_required()
def create_help_ticket():
    try:
        uid = int(get_jwt_identity())
        user = User.query.get_or_404(uid)

        data = request.get_json() or {}
        subject = (data.get("subject") or "").strip()
        category = (data.get("category") or "").strip()
        message = (data.get("message") or "").strip()

        if not subject or not message:
            return jsonify({"error": "subject and message are required"}), 400

        ticket = HelpTicket(
            user_id=user.id,
            subject=subject,
            category=category or "General",
            status="Open"
        )
        db.session.add(ticket)
        db.session.flush()

        first_message = HelpMessage(
            ticket_id=ticket.id,
            sender_id=user.id,
            sender_role=user.role,
            message=message
        )
        db.session.add(first_message)

        db.session.add(Notification(
            title="New Help Ticket",
            message=f"{user.name} created a help ticket: {subject}",
            role_target="Admin"
        ))
        db.session.add(Notification(
            title="New Help Ticket",
            message=f"{user.name} created a help ticket: {subject}",
            role_target="Staff"
        ))

        db.session.commit()

        return jsonify({
            "message": "Help ticket created successfully",
            "ticket": serialize_help_ticket(ticket, include_messages=True)
        }), 201

    except Exception as e:
        db.session.rollback()
        print("CREATE HELP TICKET ERROR:", e)
        return jsonify({"error": "Failed to create help ticket"}), 500


@api.get("/help-centre/tickets")
@jwt_required()
def list_help_tickets():
    try:
        role = get_jwt().get("role")
        uid = int(get_jwt_identity())

        if role in ["Admin", "Staff"]:
            tickets = HelpTicket.query.order_by(HelpTicket.updated_at.desc(), HelpTicket.id.desc()).all()
        else:
            tickets = HelpTicket.query.filter_by(user_id=uid).order_by(HelpTicket.updated_at.desc(), HelpTicket.id.desc()).all()

        return jsonify([serialize_help_ticket(t) for t in tickets]), 200

    except Exception as e:
        print("LIST HELP TICKETS ERROR:", e)
        return jsonify({"error": "Failed to load help tickets"}), 500


@api.get("/help-centre/tickets/<int:ticket_id>")
@jwt_required()
def get_help_ticket(ticket_id):
    try:
        role = get_jwt().get("role")
        uid = int(get_jwt_identity())

        ticket = HelpTicket.query.get_or_404(ticket_id)

        if role not in ["Admin", "Staff"] and ticket.user_id != uid:
            return jsonify({"error": "Forbidden"}), 403

        return jsonify(serialize_help_ticket(ticket, include_messages=True)), 200

    except Exception as e:
        print("GET HELP TICKET ERROR:", e)
        return jsonify({"error": "Failed to load help ticket"}), 500


@api.post("/help-centre/tickets/<int:ticket_id>/messages")
@jwt_required()
def send_help_message(ticket_id):
    try:
        role = get_jwt().get("role")
        uid = int(get_jwt_identity())

        ticket = HelpTicket.query.get_or_404(ticket_id)
        user = User.query.get_or_404(uid)

        data = request.get_json() or {}
        message = (data.get("message") or "").strip()

        if not message:
            return jsonify({"error": "message is required"}), 400

        if ticket.status == "Closed":
            return jsonify({"error": "This help ticket is closed"}), 400

        if role not in ["Admin", "Staff"]:
            return jsonify({"error": "Only Admin or Staff can reply in chat"}), 403

        new_message = HelpMessage(
            ticket_id=ticket.id,
            sender_id=user.id,
            sender_role=user.role,
            message=message
        )
        db.session.add(new_message)

        ticket.updated_at = datetime.utcnow()

        db.session.add(Notification(
            user_id=ticket.user_id,
            title="Help Centre Reply",
            message=f"{user.role} replied to your ticket: {ticket.subject}"
        ))

        db.session.commit()

        return jsonify({
            "message": "Reply sent successfully",
            "reply": serialize_help_message(new_message)
        }), 201

    except Exception as e:
        db.session.rollback()
        print("SEND HELP MESSAGE ERROR:", e)
        return jsonify({"error": "Failed to send message"}), 500


@api.put("/help-centre/tickets/<int:ticket_id>/status")
@jwt_required()
def update_help_ticket_status(ticket_id):
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Only Admin or Staff can update ticket status"}), 403

    try:
        ticket = HelpTicket.query.get_or_404(ticket_id)
        data = request.get_json() or {}
        status = (data.get("status") or "").strip()

        if status not in ["Open", "Closed"]:
            return jsonify({"error": "status must be Open or Closed"}), 400

        ticket.status = status
        ticket.updated_at = datetime.utcnow()
        db.session.commit()

        return jsonify({
            "message": "Ticket status updated successfully",
            "ticket": serialize_help_ticket(ticket)
        }), 200

    except Exception as e:
        db.session.rollback()
        print("UPDATE HELP TICKET STATUS ERROR:", e)
        return jsonify({"error": "Failed to update help ticket status"}), 500


@api.get("/complaints")
@jwt_required()
def get_complaints():
    role = get_jwt().get("role")
    uid = int(get_jwt_identity())

    q = Complaint.query

    if role not in ["Admin", "Staff"]:
        q = q.filter_by(user_id=uid)

    items = q.order_by(Complaint.id.desc()).all()

    return jsonify([
        {
            "id": c.id,
            "user_id": c.user_id,
            "type": c.type,
            "message": c.message,
            "status": c.status,
            "created_at": c.created_at.isoformat()
        }
        for c in items
    ])


@api.post("/complaints")
@jwt_required()
def add_complaint():
    try:
        uid = int(get_jwt_identity())
        data = request.get_json() or {}

        complaint_type = (data.get("type") or "").strip()
        message = (data.get("message") or "").strip()

        if not complaint_type or not message:
            return jsonify({"error": "type and message required"}), 400

        complaint = Complaint(
            user_id=uid,
            type=complaint_type,
            message=message
        )
        db.session.add(complaint)
        db.session.commit()

        return jsonify({"message": "Complaint submitted successfully"}), 201

    except Exception as e:
        db.session.rollback()
        print("COMPLAINT ERROR:", e)
        return jsonify({"error": "Failed to submit complaint"}), 500


@api.put("/complaints/<int:cid>/status")
@jwt_required()
def update_complaint_status(cid):
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        data = request.get_json() or {}
        status = (data.get("status") or "").strip()

        if status not in ["Open", "Resolved"]:
            return jsonify({"error": "status must be Open/Resolved"}), 400

        complaint = Complaint.query.get_or_404(cid)
        complaint.status = status
        db.session.commit()

        return jsonify({"message": "Complaint status updated successfully"}), 200

    except Exception as e:
        db.session.rollback()
        print("COMPLAINT STATUS ERROR:", e)
        return jsonify({"error": "Failed to update complaint status"}), 500


@api.get("/notifications")
@jwt_required()
def get_notifications():
    role = get_jwt().get("role")
    uid = int(get_jwt_identity())

    items = Notification.query.order_by(Notification.id.desc()).all()

    result = []
    for n in items:
        if n.user_id is not None:
            if n.user_id == uid:
                result.append({
                    "id": n.id,
                    "title": n.title,
                    "message": n.message,
                    "user_id": n.user_id,
                    "role_target": n.role_target,
                    "created_at": n.created_at.isoformat()
                })
        else:
            if n.role_target is None or n.role_target == role:
                result.append({
                    "id": n.id,
                    "title": n.title,
                    "message": n.message,
                    "user_id": n.user_id,
                    "role_target": n.role_target,
                    "created_at": n.created_at.isoformat()
                })

    return jsonify(result)


@api.post("/notifications")
@jwt_required()
def add_notification():
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        data = request.get_json() or {}

        title = (data.get("title") or "").strip()
        message = (data.get("message") or "").strip()
        role_target = data.get("role_target")
        user_id = data.get("user_id")

        if not title or not message:
            return jsonify({"error": "title and message required"}), 400

        if role_target not in [None, "Admin", "User", "Staff"]:
            return jsonify({"error": "role_target invalid"}), 400

        notification = Notification(
            user_id=user_id,
            title=title,
            message=message,
            role_target=role_target
        )
        db.session.add(notification)
        db.session.commit()

        return jsonify({"message": "Notification created successfully"}), 201

    except Exception as e:
        db.session.rollback()
        print("NOTIFICATION ERROR:", e)
        return jsonify({"error": "Failed to create notification"}), 500


@api.post("/billing/create")
@jwt_required()
def create_bill():
    if not require_roles("Admin")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        data = request.get_json() or {}

        user_id = data.get("user_id")
        amount = data.get("amount")
        period = (data.get("period") or "").strip()
        bill_type = (data.get("bill_type") or "monthly").strip().lower()

        if not user_id or amount is None or not period:
            return jsonify({"error": "user_id, amount and period are required"}), 400

        if bill_type not in ["monthly", "daily"]:
            return jsonify({"error": "bill_type must be monthly or daily"}), 400

        try:
            user_id = int(user_id)
            amount = float(amount)
            if bill_type == "daily":
                period = normalize_to_iso_date(period)
        except Exception:
            return jsonify({"error": "user_id and amount must be valid numbers"}), 400

        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        existing = Bill.query.filter_by(
            user_id=user_id,
            month=period,
            bill_type=bill_type,
            parent_bill_id=None
        ).first()

        if existing:
            return jsonify({"error": "Bill already exists for this user, type and period"}), 400

        bill = Bill(
            user_id=user_id,
            month=period,
            bill_type=bill_type,
            meal_type=None,
            attendance_id=None,
            parent_bill_id=None,
            amount=amount,
            status="Unpaid"
        )
        db.session.add(bill)
        db.session.flush()

        message = f"Your {bill_type.capitalize()} bill for {period} of ₹{amount} has been generated."
        notif = Notification(
            user_id=user.id,
            title="Bill Generated",
            message=message
        )
        db.session.add(notif)

        db.session.commit()

        if user.email:
            try:
                send_bill_email(
                    user.email,
                    user.name,
                    bill_type.capitalize(),
                    period,
                    amount
                )
            except Exception as mail_error:
                print("BILL EMAIL ERROR:", mail_error)

        return jsonify({
            "message": "Bill created successfully",
            "bill": serialize_bill_row(bill, include_user=True)
        }), 201

    except Exception as e:
        db.session.rollback()
        print("BILL CREATE ERROR:", e)
        return jsonify({"error": "Bill creation failed"}), 500


@api.post("/billing/generate-monthly")
@jwt_required()
def generate_monthly_bill_from_attendance():
    try:
        data = request.get_json() or {}
        period = (data.get("period") or datetime.utcnow().strftime("%Y-%m")).strip()
        role = get_jwt().get("role")
        logged_in_user_id = int(get_jwt_identity())

        try:
            datetime.strptime(period, "%Y-%m")
        except Exception:
            return jsonify({"error": "period must be in YYYY-MM format"}), 400

        target_user_id = logged_in_user_id
        if role in ["Admin", "Staff"] and data.get("user_id") is not None:
            try:
                target_user_id = int(data.get("user_id"))
            except Exception:
                return jsonify({"error": "user_id must be a valid number"}), 400

        user = User.query.get(target_user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        existing_monthly = Bill.query.filter_by(
            user_id=target_user_id,
            month=period,
            bill_type="monthly",
            meal_type="Monthly Attendance",
            parent_bill_id=None
        ).first()
        if existing_monthly:
            return jsonify({"error": "Monthly attendance bill already exists for this month"}), 400

        source_bills = get_unmerged_daily_bills_for_month(target_user_id, period)
        if not source_bills:
            return jsonify({"error": "No unpaid daily attendance bills found for this month"}), 400

        total_amount = round(sum(float(b.amount or 0) for b in source_bills), 2)
        monthly_bill = Bill(
            user_id=target_user_id,
            month=period,
            bill_type="monthly",
            meal_type="Monthly Attendance",
            attendance_id=None,
            parent_bill_id=None,
            amount=total_amount,
            status="Unpaid"
        )
        db.session.add(monthly_bill)
        db.session.flush()

        for bill_row in source_bills:
            bill_row.parent_bill_id = monthly_bill.id
            bill_row.status = "Merged"

        db.session.add(Notification(
            user_id=user.id,
            title="Monthly Attendance Bill Generated",
            message=f"Your monthly attendance bill for {period} has been generated: ₹{total_amount}"
        ))

        db.session.commit()

        if user.email:
            try:
                send_bill_email(user.email, user.name, "Monthly Attendance", period, total_amount)
            except Exception as mail_error:
                print("MONTHLY ATTENDANCE EMAIL ERROR:", mail_error)

        return jsonify({
            "message": "Monthly attendance bill generated successfully",
            "bill": serialize_bill_row(monthly_bill, include_user=True),
            "source_bill_ids": [b.id for b in source_bills]
        }), 201

    except Exception as e:
        db.session.rollback()
        print("GENERATE MONTHLY BILL ERROR:", e)
        return jsonify({"error": "Failed to generate monthly attendance bill"}), 500


@api.get("/billing/my")
@jwt_required()
def my_bills():
    try:
        user_id = int(get_jwt_identity())
        bills = Bill.query.filter_by(user_id=user_id).order_by(Bill.id.desc()).all()
        visible_bills = [b for b in bills if b.parent_bill_id is None]
        return jsonify([serialize_bill_row(b) for b in visible_bills]), 200
    except Exception as e:
        print("MY BILLS ERROR:", e)
        return jsonify({"error": "Failed to load my bills"}), 500


@api.get("/billing/all")
@jwt_required()
def all_bills():
    if not require_roles("Admin", "Staff")():
        return jsonify({"error": "Forbidden"}), 403

    try:
        bills = Bill.query.order_by(Bill.id.desc()).all()
        visible_bills = [b for b in bills if b.parent_bill_id is None]
        return jsonify([serialize_bill_row(b, include_user=True) for b in visible_bills]), 200
    except Exception as e:
        print("ALL BILLS ERROR:", e)
        return jsonify({"error": "Failed to load all bills"}), 500


@api.post("/billing/pay")
@jwt_required()
def pay_bill():
    try:
        user_id = int(get_jwt_identity())

        bill_id = request.form.get("bill_id")
        mode = (request.form.get("mode") or "UPI").strip()
        note = (request.form.get("note") or "").strip()
        proof = request.files.get("proof")

        if not bill_id:
            return jsonify({"error": "bill_id is required"}), 400

        try:
            bill_id = int(bill_id)
        except Exception:
            return jsonify({"error": "Invalid bill_id"}), 400

        bill = Bill.query.get_or_404(bill_id)

        role = get_jwt().get("role")
        if role not in ["Admin", "Staff"] and bill.user_id != user_id:
            return jsonify({"error": "You can pay only your own bill"}), 403

        if bill.parent_bill_id:
            return jsonify({"error": "This bill is already included in a monthly bill"}), 400

        if bill.status == "Paid":
            return jsonify({"error": "Bill is already paid"}), 400

        proof_filename = None
        if proof:
            upload_dir = os.path.join(os.getcwd(), "uploads")
            os.makedirs(upload_dir, exist_ok=True)
            original_name = secure_filename(proof.filename or "proof")
            timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S")
            proof_filename = f"{timestamp}_{original_name}"
            proof.save(os.path.join(upload_dir, proof_filename))

        receipt_no = f"RCPT-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{bill.id}"
        payment = Payment(
            bill_id=bill.id,
            mode=mode,
            receipt_no=receipt_no,
            proof_filename=proof_filename,
            note=note
        )

        bill.status = "Paid"
        db.session.add(payment)

        if bill.bill_type == "monthly":
            child_bills = Bill.query.filter_by(parent_bill_id=bill.id).all()
            for child in child_bills:
                child.status = "Paid"

        db.session.commit()

        return jsonify({
            "message": "Payment recorded successfully",
            "receipt": {
                "receipt_no": receipt_no,
                "bill_id": bill.id,
                "bill_type": bill.bill_type,
                "meal_type": bill.meal_type,
                "period": bill.month,
                "amount": bill.amount,
                "mode": mode,
                "paid_at": payment.paid_at.isoformat() if payment.paid_at else datetime.utcnow().isoformat(),
                "proof_url": f"http://127.0.0.1:5000/api/uploads/{proof_filename}" if proof_filename else None
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        print("PAY BILL ERROR:", e)
        return jsonify({"error": "Failed to record payment"}), 500


@api.get("/billing/receipt/<int:bill_id>")
@jwt_required()
def billing_receipt(bill_id):
    try:
        user_id = int(get_jwt_identity())
        role = get_jwt().get("role")

        bill = Bill.query.get_or_404(bill_id)
        if role not in ["Admin", "Staff"] and bill.user_id != user_id:
            return jsonify({"error": "Forbidden"}), 403

        payment = Payment.query.filter_by(bill_id=bill.id).order_by(Payment.id.desc()).first()
        user = User.query.get(bill.user_id)

        return jsonify({
            "bill_id": bill.id,
            "bill_type": bill.bill_type,
            "meal_type": bill.meal_type,
            "period": bill.month,
            "amount": bill.amount,
            "status": bill.status,
            "user": {
                "id": user.id if user else None,
                "name": user.name if user else "Unknown",
                "email": user.email if user else "Unknown"
            },
            "payment": {
                "mode": payment.mode if payment else None,
                "receipt_no": payment.receipt_no if payment else None,
                "paid_at": payment.paid_at.isoformat() if payment and payment.paid_at else None,
                "proof_url": f"http://127.0.0.1:5000/api/uploads/{payment.proof_filename}" if payment and payment.proof_filename else None,
                "note": payment.note if payment else None
            } if payment else None
        }), 200

    except Exception as e:
        print("RECEIPT ERROR:", e)
        return jsonify({"error": "Failed to load receipt"}), 500
